#! python3
#-*- encoding: utf-8 -*-
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet
# Usage:
#
# Author : qmeng
# MailTo : qmeng1128@163.com
# QQ     : 1163306125
# Blog   : http://blog.csdn.net/Mq_Go/
# Create : 2018-02-15 15:56:52
# Version: 1.0
#
import openpyxl
import smtplib
import sys
from smtplib import SMTP_SSL
from email import encoders
from email.header import Header
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr

def _format_addr(s):
    name, addr = parseaddr(s)
    return formataddr((Header(name, 'utf-8').encode(), addr))
#打开Excel文件
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1,column=lastCol).value

#找到未付款成员
unpaidMembers = {}
for r in range(2,sheet.max_row+1):
	payment = sheet.cell(row=r,column=lastCol).value
	if payment != 'paid':
		name  = sheet.cell(row=r,column=1).value
		email = sheet.cell(row=r,column=2).value
		unpaidMembers[name] = email

#发送定制的电子邮件提醒
mailInfo = {}
mailInfo['Username'] = input('Username:')
mailInfo['Password'] = input('Password:')
mailInfo['Hostname'] = input('Hostname:')

msg = MIMEText('hello,这是来自Python的信\n 新年快乐！！！','plain','utf-8')
msg['From'] =  _format_addr('奇奇 <%s>' % mailInfo['Username'] )
msg['Subject'] = Header('来自新年的问候...','utf-8').encode()

for name,to in unpaidMembers.items():
	msg['To'] = _format_addr('用户%s <%s>' % (name,to))
	smtp = SMTP_SSL(mailInfo["Hostname"])
	smtp.set_debuglevel(1)
	smtp.ehlo(mailInfo["Hostname"])
	smtp.login(mailInfo["Username"],mailInfo["Password"])
	smtp.sendmail(mailInfo['Username'],to,msg.as_string())
	smtp.quit()
	print('向用户 '+ name + ' ' + to +' 发送定制的电子邮件...成功')

		
